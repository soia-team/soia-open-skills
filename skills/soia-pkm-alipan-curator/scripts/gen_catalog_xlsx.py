#!/usr/bin/env python3
"""Incrementally generate a master Excel catalog and per-partition detail workbooks."""

from __future__ import annotations

import argparse
import importlib.util
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path

from catalog_xlsx.cache import (
    commit_manifest,
    default_cache_dir,
    normalize_release_metadata,
    prepare_incremental,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BUILDER = SCRIPT_DIR / "catalog_xlsx" / "build_workbooks.mjs"
MASTER_FILENAME = "00_阿里云盘馆藏总索引.xlsx"
OUTPUT_ENV_NAME = "ALIPAN_CURATOR_OUTPUT_DIR"
CONFIG_FILE_ENV_NAME = "SOIA_PKM_ALIPAN_CURATOR_CONFIG_FILE"


def default_config_file() -> Path:
    if os.name == "nt":
        base = Path(os.environ.get("APPDATA") or (Path.home() / "AppData" / "Roaming"))
    else:
        base = Path.home() / ".config"
    return base / "soia-skills" / "soia-open-skills" / "soia-pkm" / "soia-pkm-alipan-curator" / "config.yml"


def parse_config_scalar(value: str) -> str:
    value = value.strip()
    if not value or value in {"null", "~"}:
        return ""
    if value[0] in {"'", '"'} and value[-1:] == value[0]:
        return value[1:-1]
    return value.split(" #", 1)[0].strip()


def configured_output_dir() -> str | None:
    config_path = Path(os.environ[CONFIG_FILE_ENV_NAME]).expanduser() if os.environ.get(CONFIG_FILE_ENV_NAME) else default_config_file()
    if not config_path.is_file():
        return None
    try:
        in_env = False
        for raw_line in config_path.read_text(encoding="utf-8").splitlines():
            stripped = raw_line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            indent = len(raw_line) - len(raw_line.lstrip(" "))
            if indent == 0:
                in_env = stripped == "env:"
                continue
            if not in_env or indent < 2 or ":" not in stripped:
                continue
            key, value = stripped.split(":", 1)
            if key.strip() == OUTPUT_ENV_NAME:
                return parse_config_scalar(value)
    except OSError as error:
        print(f"⚠️ 无法读取私有 config.yml {config_path}: {error}", file=sys.stderr)
    return None


def resolve_output_dir(cli_output_dir: Path | None) -> tuple[Path, str]:
    value: str | Path | None = cli_output_dir
    source = "cli"
    if value is None:
        value = os.environ.get(OUTPUT_ENV_NAME)
        source = "environment"
    if not value:
        value = configured_output_dir()
        source = "config"
    if not value:
        value = Path.home() / "Downloads" / "soia-pkm-alipan-curator"
        source = "default"
        print(
            f"输出到默认目录 {value}（可用 --output-dir 或 config ALIPAN_CURATOR_OUTPUT_DIR 覆盖）",
            file=sys.stderr,
        )
    resolved = Path(os.path.expandvars(str(value))).expanduser()
    if not resolved.is_absolute():
        raise SystemExit("输出目录必须是绝对路径；不要使用 cwd 相对路径。")
    resolved = resolved.resolve()
    resolved.mkdir(parents=True, exist_ok=True)
    return resolved, source


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="把 OB 云盘馆藏 Markdown 增量生成成轻量总索引 + 分区明细 Excel。"
    )
    parser.add_argument("--catalog", type=Path, required=True, help="馆藏总览 Markdown 文件")
    parser.add_argument("--search-dir", type=Path, required=True, help="分区全文检索 Markdown 目录")
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="用户交付物目录；未提供时按环境变量、私有 config.yml、默认 Downloads 目录回退",
    )
    parser.add_argument("--cache-dir", type=Path, help="增量缓存目录；默认按数据源路径生成用户级缓存")
    parser.add_argument("--node", default=os.environ.get("SOIA_ARTIFACT_NODE", "node"), help="Node.js 可执行文件")
    parser.add_argument(
        "--artifact-runtime",
        type=Path,
        default=os.environ.get("SOIA_ARTIFACT_RUNTIME"),
        help="包含 node_modules/@oai/artifact-tool 的运行目录",
    )
    parser.add_argument(
        "--soffice",
        default=os.environ.get("SOIA_SOFFICE"),
        help="可选：LibreOffice/soffice 路径，用于预计算 HYPERLINK 显示值；openpyxl 兜底路径下还用于公式错误扫描",
    )
    parser.add_argument(
        "--renderer",
        choices=["auto", "artifact-tool", "openpyxl"],
        default=os.environ.get("SOIA_CATALOG_XLSX_RENDERER", "auto"),
        help=(
            "Excel 作者层实现：artifact-tool 需要宿主提供的 @oai/artifact-tool 运行时（格式更丰富，"
            "支持公式错误扫描与截图预览）；openpyxl 是不依赖特定平台的通用兜底，任何装有 openpyxl 的"
            "环境都能跑；auto（默认）优先用可用的 artifact-tool，缺失时自动退回 openpyxl，保证换一个"
            "没有 artifact-tool 访问权限的 AI/宿主也能产出有效索引。"
        ),
    )
    parser.add_argument("--force", action="store_true", help="忽略缓存，重建全部分区")
    parser.add_argument("--verify", action="store_true", help="渲染总索引与本次变化分区的全部工作表")
    parser.add_argument(
        "--release-metadata",
        help="可选：release metadata JSON 字符串或 JSON 文件（五项字段均需非空，时间须带时区）",
    )
    parser.add_argument("--json", action="store_true", help="只输出 JSON 结果")
    return parser.parse_args()


def load_release_metadata(argument: str | None) -> dict[str, str] | None:
    if argument is None:
        return None
    raw = argument.strip()
    if not raw:
        raise ValueError("--release-metadata 不能为空")
    if raw.startswith("{"):
        source = raw
    else:
        path = Path(raw).expanduser()
        if not path.is_file():
            raise ValueError("--release-metadata 必须是 JSON 对象或存在的 JSON 文件")
        source = path.read_text(encoding="utf-8")
    try:
        return normalize_release_metadata(json.loads(source))
    except json.JSONDecodeError as error:
        raise ValueError("--release-metadata 不是有效 JSON") from error


def resolve_node(node_arg: str) -> str | None:
    node = shutil.which(node_arg) if os.sep not in node_arg else node_arg
    if not node or not Path(node).exists():
        return None
    return str(Path(node).resolve())


def openpyxl_available() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


def resolve_renderer(args: argparse.Namespace) -> str:
    """Pick the Excel author backend. artifact-tool is an optional, platform-specific
    fast path (richer styling, built-in formula-error scan + screenshot QA); openpyxl
    is the platform-agnostic fallback so any AI/host without artifact-tool access can
    still produce a fully usable catalog Excel. See references/catalog-excel.md。
    """
    node_path = resolve_node(args.node)
    artifact_runtime = Path(args.artifact_runtime).expanduser().resolve() if args.artifact_runtime else None
    artifact_package = (artifact_runtime / "node_modules" / "@oai" / "artifact-tool") if artifact_runtime else None
    artifact_ok = bool(node_path and artifact_runtime and artifact_package and artifact_package.exists())

    if args.renderer == "artifact-tool":
        if not artifact_runtime:
            raise SystemExit(
                "缺少 --artifact-runtime：请先让宿主加载 spreadsheet workspace dependencies，"
                "在临时工作目录创建 node_modules 软链后传入该目录。也可改用 --renderer openpyxl 或"
                "--renderer auto 使用不依赖平台的兜底路径。"
            )
        if not node_path:
            raise SystemExit(f"Node.js 不可用：{args.node}")
        if not artifact_package.exists():
            raise SystemExit(f"artifact runtime 中未找到 @oai/artifact-tool：{artifact_package}")
        args.node = node_path
        args.artifact_runtime = artifact_runtime
        return "artifact-tool"

    if args.renderer == "openpyxl":
        if not openpyxl_available():
            raise SystemExit(
                "openpyxl 不可用：请先安装（pip install openpyxl 或 uv pip install openpyxl），"
                "或改用 --renderer artifact-tool 并提供有效的 --artifact-runtime。"
            )
        return "openpyxl"

    # auto：优先复用可用的 artifact-tool（格式更丰富），缺失时自动退回 openpyxl 兜底。
    if artifact_ok:
        args.node = node_path
        args.artifact_runtime = artifact_runtime
        return "artifact-tool"
    if openpyxl_available():
        return "openpyxl"
    raise SystemExit(
        "既没有可用的 --artifact-runtime（@oai/artifact-tool 运行时），也没有安装 openpyxl 兜底依赖；"
        "请安装 openpyxl（pip install openpyxl）或提供有效的 --artifact-runtime 后重试。"
    )


def validate_inputs(args: argparse.Namespace) -> None:
    args.output_dir, _ = resolve_output_dir(args.output_dir)
    args.output = args.output_dir / MASTER_FILENAME
    if not args.catalog.is_file():
        raise SystemExit(f"catalog 不存在：{args.catalog}")
    if not args.search_dir.is_dir():
        raise SystemExit(f"search-dir 不存在：{args.search_dir}")
    if args.output.suffix.lower() != ".xlsx":
        raise SystemExit("output 必须以 .xlsx 结尾")
    args.renderer = resolve_renderer(args)
    if args.soffice:
        soffice = shutil.which(args.soffice) if os.sep not in args.soffice else args.soffice
        if not soffice or not Path(soffice).exists():
            raise SystemExit(f"soffice 不可用：{args.soffice}")
        args.soffice = str(Path(soffice).resolve())


def recalculate_with_soffice(soffice: str, outputs: list[Path], cache_dir: Path) -> None:
    for output in outputs:
        if not output.exists():
            raise RuntimeError(f"待重算工作簿不存在：{output}")
        with tempfile.TemporaryDirectory(prefix="recalc-", dir=cache_dir) as temporary:
            command = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", temporary, str(output)]
            process = subprocess.run(command, text=True, capture_output=True, check=False)
            if process.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice 重算失败：{output}\nstdout={process.stdout}\nstderr={process.stderr}"
                )
            recalculated = Path(temporary) / output.name
            if not recalculated.exists():
                raise RuntimeError(f"LibreOffice 未生成重算文件：{recalculated}")
            recalculated.replace(output)


EXCEL_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def scan_formula_errors(outputs: list[Path]) -> list[dict]:
    """openpyxl 兜底路径下的公式错误扫描（对齐 artifact-tool 的 workbook.inspect()）。

    openpyxl 本身不计算公式；只能扫描已缓存的计算值，因此只有在 recalculate_with_soffice()
    先用 LibreOffice 重算过一遍之后调用才有意义。
    """
    from openpyxl import load_workbook

    errors: list[dict] = []
    for output in outputs:
        workbook = load_workbook(output, data_only=True, read_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.strip() in EXCEL_ERROR_TOKENS:
                            errors.append(
                                {"file": str(output), "sheet": sheet.title, "cell": cell.coordinate, "value": cell.value}
                            )
        finally:
            workbook.close()
    return errors


def cleanup_inspection_sidecars(outputs: list[Path]) -> None:
    """清理 artifact-tool inspect 生成的调试文件，只保留正式 xlsx。"""
    for output in outputs:
        Path(f"{output}.inspect.ndjson").unlink(missing_ok=True)


def cleanup_stale_partition_outputs(plan: dict) -> list[str]:
    """成功提交新 manifest 后，移除已从检索源消失的旧分区产物。"""
    detail_dir = Path(plan["detailDir"]).resolve()
    partition_cache_dir = Path(plan["cacheDir"]).resolve() / "partitions"
    managed = {Path(item["output"]).resolve() for item in plan["partitions"]}
    candidates = {item.resolve() for item in detail_dir.glob("*.xlsx") if item.resolve() not in managed}
    candidates.update(Path(item["output"]).resolve() for item in plan.get("stalePartitions", []))
    removed = []
    for output in sorted(candidates):
        if output.parent != detail_dir or output.suffix.lower() != ".xlsx":
            raise RuntimeError(f"拒绝清理分区明细目录之外的旧产物：{output}")
        Path(f"{output}.inspect.ndjson").unlink(missing_ok=True)
        if output.exists():
            output.unlink()
            removed.append(str(output))
    for item in plan.get("stalePartitions", []):
        cache = Path(item["cache"]).resolve()
        if cache.parent == partition_cache_dir and cache.suffix.lower() == ".json":
            cache.unlink(missing_ok=True)
    return removed


def run() -> dict:
    args = parse_args()
    validate_inputs(args)
    release_metadata = load_release_metadata(args.release_metadata)
    started = time.monotonic()
    cache_dir = (args.cache_dir or default_cache_dir(args.catalog, args.search_dir)).resolve()
    plan = prepare_incremental(
        catalog_path=args.catalog,
        search_dir=args.search_dir,
        output_path=args.output,
        cache_dir=cache_dir,
        force=args.force,
        verify=args.verify,
        release_metadata=release_metadata,
    )
    changed = plan["changedPartitions"]
    managed_outputs = [Path(plan["outputPath"])]
    managed_outputs.extend(Path(item["output"]) for item in plan["partitions"])
    outputs = []
    removed_outputs = []
    if plan["buildMaster"]:
        outputs.append(Path(plan["outputPath"]))
    outputs.extend(Path(item["output"]) for item in plan["partitions"] if item["changed"])

    qa = {"renderer": args.renderer}
    if outputs:
        if args.renderer == "artifact-tool":
            environment = os.environ.copy()
            node_options = os.environ.get("SOIA_ARTIFACT_NODE_OPTIONS")
            if node_options:
                environment["NODE_OPTIONS"] = node_options
            command = [
                args.node,
                str(BUILDER),
                "--plan",
                str(cache_dir / "build-plan.json"),
                "--artifact-runtime",
                str(args.artifact_runtime),
            ]
            process = subprocess.run(command, text=True, capture_output=True, env=environment, check=False)
            if process.returncode != 0:
                raise RuntimeError(
                    f"artifact-tool 生成失败（缓存未提交，下次会继续重建）：\n"
                    f"stdout={process.stdout}\nstderr={process.stderr}"
                )
        else:
            from catalog_xlsx.build_workbooks_fallback import build as build_with_openpyxl

            try:
                build_with_openpyxl(cache_dir / "build-plan.json")
            except Exception as error:
                raise RuntimeError(f"openpyxl 兜底生成失败（缓存未提交，下次会继续重建）：{error}") from error
        if args.soffice:
            recalculate_with_soffice(args.soffice, outputs, cache_dir)
        commit_manifest(plan)
    removed_outputs = cleanup_stale_partition_outputs(plan)
    cleanup_inspection_sidecars(managed_outputs)

    if args.verify:
        if args.renderer == "artifact-tool":
            qa["note"] = "公式错误扫描与截图预览已由 artifact-tool 内置完成，见各输出旁的调试信息。"
        elif args.soffice and outputs:
            formula_errors = scan_formula_errors(outputs)
            qa["formulaErrorCount"] = len(formula_errors)
            qa["formulaErrors"] = formula_errors[:20]
            qa["previewsAvailable"] = False
        else:
            qa["formulaErrorCount"] = None
            qa["previewsAvailable"] = False
            qa["note"] = (
                "openpyxl 兜底模式未提供 --soffice：跳过公式错误扫描与截图预览，公式仅在 "
                "Excel/WPS/LibreOffice 打开时才会实际计算；交付前建议人工打开核对，或补充 --soffice。"
            )

    payload = {
        "status": "updated" if outputs else "unchanged",
        "renderer": args.renderer,
        "qa": qa if args.verify else None,
        "master": plan["outputPath"],
        "detailDir": plan["detailDir"],
        "changedPartitions": changed,
        "rebuilt": [str(item) for item in outputs],
        "reusedPartitions": [item["partition"] for item in plan["partitions"] if not item["changed"]],
        "removedPartitions": [item["partition"] for item in plan.get("stalePartitions", [])],
        "removedOutputs": removed_outputs,
        "cacheDir": str(cache_dir),
        "elapsedSeconds": round(time.monotonic() - started, 3),
        "verified": args.verify,
        "recalculated": bool(args.soffice and outputs),
        "releaseMetadata": release_metadata,
    }
    if args.json:
        print(json.dumps(payload, ensure_ascii=False))
    else:
        print(f"status={payload['status']}")
        print(f"renderer={payload['renderer']}")
        print(f"master={payload['master']}")
        print(f"rebuilt={len(payload['rebuilt'])}")
        print(f"changed_partitions={len(changed)}")
        print(f"reused_partitions={len(payload['reusedPartitions'])}")
        print(f"elapsed_seconds={payload['elapsedSeconds']}")
    return payload


if __name__ == "__main__":
    try:
        run()
    except (ValueError, RuntimeError) as error:
        print(f"error: {error}", file=sys.stderr)
        raise SystemExit(1) from error
