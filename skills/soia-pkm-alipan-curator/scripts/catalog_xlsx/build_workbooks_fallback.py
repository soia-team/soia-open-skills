"""openpyxl fallback renderer for the master catalog and per-partition detail workbooks.

Used when the host does not expose the ``@oai/artifact-tool`` spreadsheet runtime
(``build_workbooks.mjs``'s dependency). Any Python environment with ``openpyxl``
installed can run this renderer, so hosts/AIs without access to that bespoke runtime
still produce a fully usable catalog Excel: same sheet names, same headers, same data,
same cross-sheet formulas as the artifact-tool path.

Known gaps versus the artifact-tool path (documented, not silently dropped):
- No automatic formula-error scan (artifact-tool's ``workbook.inspect()``); formulas are
  only evaluated when the workbook is opened in Excel/WPS/LibreOffice. ``gen_catalog_xlsx.py``
  can still run a best-effort scan via ``--soffice`` (see ``scan_formula_errors_via_soffice``
  in that script).
- No automatic screenshot QA previews (artifact-tool's ``workbook.render()``).
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


MB = 1024**2
GB = 1024**3

TYPE_NOTES = {
    "视频": "常见视频容器",
    "音频": "音乐、课程音频与有声书",
    "电子书": "PDF/EPUB/MOBI 等",
    "Office文档": "Word/Excel/PPT/CSV",
    "文本与网页": "TXT/MD/HTML/JSON 等",
    "图片": "常见位图与矢量图",
    "压缩包": "ZIP/RAR/7Z/ISO 等",
    "字幕": "SRT/ASS/VTT 等",
    "软件与安装包": "EXE/DMG/APK/DLL 等",
    "代码与数据": "源码、脚本、数据库与数据文件",
    "其他": "未命中已知扩展名规则",
}

COLORS = {
    "navy": "16324F",
    "teal": "0F766E",
    "pale": "E8F3F1",
    "amber": "FEF3C7",
    "light_gray": "E2E8F0",
    "white": "FFFFFF",
}

_THIN_TEAL = Side(style="thin", color=COLORS["teal"])
_HEADER_BORDER = Border(top=_THIN_TEAL, bottom=_THIN_TEAL, left=_THIN_TEAL, right=_THIN_TEAL)
_LINK_FONT = Font(color="0563C1")


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def _range_bounds(cell_range: str) -> tuple[int, int, int, int]:
    from openpyxl.utils import range_boundaries

    return range_boundaries(cell_range)


def col(index0: int) -> str:
    """0-indexed column number -> Excel column letter (mirrors JS excelColumn())."""
    return get_column_letter(index0 + 1)


def excel_text(value: Any) -> str:
    return str(value if value is not None else "").replace('"', '""')


def prepare_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False


def title_band(sheet: Worksheet, cell_range: str, title: str) -> None:
    sheet.merge_cells(cell_range)
    left, top, _, _ = _range_bounds(cell_range)
    cell = sheet.cell(row=top, column=left)
    cell.value = title
    cell.fill = _fill(COLORS["navy"])
    cell.font = Font(bold=True, color=COLORS["white"], size=16)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[top].height = 30


def note_band(sheet: Worksheet, cell_range: str, text: str, warning: bool = False) -> None:
    sheet.merge_cells(cell_range)
    left, top, _, _ = _range_bounds(cell_range)
    cell = sheet.cell(row=top, column=left)
    cell.value = text
    cell.fill = _fill(COLORS["amber"] if warning else COLORS["pale"])
    cell.font = Font(color="92400E" if warning else COLORS["navy"], italic=warning)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[top].height = 38 if warning else 30


def style_header(sheet: Worksheet, cell_range: str) -> None:
    left, top, right, bottom = _range_bounds(cell_range)
    for row in range(top, bottom + 1):
        for column in range(left, right + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = _fill(COLORS["teal"])
            cell.font = Font(bold=True, color=COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _HEADER_BORDER
    sheet.row_dimensions[top].height = 28


def set_column_widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def add_table(sheet: Worksheet, cell_range: str, name: str) -> None:
    table = Table(displayName=name, ref=cell_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False
    )
    sheet.add_table(table)


def friendly_local_path(master_path: str, detail_path: str) -> str:
    base = master_path.replace("\\", "/").rsplit("/", 1)[0]
    target = detail_path.replace("\\", "/")
    prefix = f"{base}/"
    return target[len(prefix):] if target.startswith(prefix) else target


def release_metadata_rows(release_metadata: dict[str, str]) -> list[list[str]]:
    return [
        ["catalog_release_id", release_metadata["catalog_release_id"]],
        ["index_updated_at", release_metadata["index_updated_at"]],
        ["snapshot_at", release_metadata["snapshot_at"]],
        ["catalog_schema_version", release_metadata["catalog_schema_version"]],
        ["source_fingerprint", release_metadata["source_fingerprint"]],
    ]


_TYPE_ORDER = list(TYPE_NOTES)


def _type_sort_key(type_label: str) -> tuple[int, str]:
    # Canonical TYPE_RULES order, not locale collation (Python's plain sorted() and
    # JS's localeCompare("zh-CN") don't agree on Chinese ordering; both renderers use
    # this same fixed list so the two backends produce identical row order).
    try:
        return (_TYPE_ORDER.index(type_label), type_label)
    except ValueError:
        return (len(_TYPE_ORDER), type_label)


def type_stats_from_files(files: list[dict[str, Any]]) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, Any]] = {}
    for row in files:
        entry = stats.setdefault(row["type"], {"type": row["type"], "count": 0, "bytes": 0})
        entry["count"] += 1
        entry["bytes"] += row["sizeBytes"]
    return sorted(stats.values(), key=lambda item: _type_sort_key(item["type"]))


def extension_stats_from_files(files: list[dict[str, Any]]) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, Any]] = {}
    for row in files:
        entry = stats.setdefault(
            row["ext"], {"ext": row["ext"], "type": row["type"], "count": 0, "bytes": 0}
        )
        entry["count"] += 1
        entry["bytes"] += row["sizeBytes"]
    return sorted(stats.values(), key=lambda item: (-item["count"], item["ext"]))


# openpyxl auto-promotes any string cell value starting with "=" into a live formula
# (data_type "f"). Alipan file/folder names are attacker-controlled (anyone who can
# upload a file can name it "=cmd|'/c calc'!A0"), so every raw data cell must be forced
# back to a literal string after assignment -- otherwise the name silently becomes an
# executable formula the moment the workbook is opened in Excel/WPS. Also guards +/-/@,
# the other classic CSV/formula-injection trigger characters, for defense in depth.
_FORMULA_TRIGGER_PREFIXES = ("=", "+", "-", "@")


def _write_row(sheet: Worksheet, row: int, start_col0: int, values: list[Any]) -> None:
    for offset, value in enumerate(values):
        if value is not None:
            cell = sheet.cell(row=row, column=start_col0 + 1 + offset)
            cell.value = value
            if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_PREFIXES):
                cell.data_type = "s"


def _number_format(sheet: Worksheet, cell_range: str, fmt: str) -> None:
    left, top, right, bottom = _range_bounds(cell_range)
    for row in range(top, bottom + 1):
        for column in range(left, right + 1):
            sheet.cell(row=row, column=column).number_format = fmt


def _font(sheet: Worksheet, cell_range: str, font: Font) -> None:
    left, top, right, bottom = _range_bounds(cell_range)
    for row in range(top, bottom + 1):
        for column in range(left, right + 1):
            sheet.cell(row=row, column=column).font = font


def compute_generated_at(plan: dict[str, Any]) -> str:
    release_metadata = plan.get("releaseMetadata")
    if release_metadata and release_metadata.get("index_updated_at"):
        return release_metadata["index_updated_at"]
    configured_tz = os.environ.get("SOIA_CATALOG_TIME_ZONE", "").strip()
    now = datetime.now()
    if configured_tz:
        from zoneinfo import ZoneInfo

        now = datetime.now(ZoneInfo(configured_tz))
    return now.strftime("%Y-%m-%d %H:%M")


def build_master_workbook(aggregate: dict[str, Any], plan: dict[str, Any], generated_at: str) -> tuple[Workbook, dict[str, Any]]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    usage = workbook.create_sheet("00_使用说明")
    directories_sheet = workbook.create_sheet("01_目录索引")
    entry_sheet = workbook.create_sheet("02_明细入口")
    type_sheet = workbook.create_sheet("03_类型统计")
    partition_sheet = workbook.create_sheet("04_分区统计")
    extension_sheet = workbook.create_sheet("05_扩展名统计")
    for sheet in (usage, directories_sheet, entry_sheet, type_sheet, partition_sheet, extension_sheet):
        prepare_sheet(sheet)

    catalog = aggregate["catalog"]
    directories = aggregate["directories"]
    type_stats = aggregate["typeStats"]
    extension_stats = aggregate["extensionStats"]
    partition_stats = aggregate["partitionStats"]
    indexed_files = aggregate["indexedFiles"]
    indexed_bytes = aggregate["indexedBytes"]
    partition_end = len(partition_stats) + 4

    title_band(usage, "A1:J1", "阿里云盘馆藏总索引")
    note_band(
        usage,
        "A2:J2",
        "轻量总入口：目录与统计保留在本工作簿，单文件明细按分区拆到相邻工作簿。以后只重建发生变化的分区，不再全量生成 6.8 万行。",
    )
    _write_row(usage, 3, 0, ["全盘文件", "", "已展开明细", "", "明细覆盖率", "", "目录索引行", "", "明细估算体量(GB)", ""])
    style_header(usage, "A3:J3")
    _write_row(usage, 4, 0, [catalog["totalFiles"]])
    usage.cell(row=4, column=3).value = f"=SUM('02_明细入口'!$D$5:$D${partition_end})"
    usage.cell(row=4, column=5).value = "=C4/A4"
    usage.cell(row=4, column=7).value = f"=COUNTA('01_目录索引'!$A$5:$A${len(directories) + 4})"
    usage.cell(row=4, column=9).value = f"=SUM('02_明细入口'!$F$5:$F${partition_end})/$B$23"
    for merge_range in ("A4:B4", "C4:D4", "E4:F4", "G4:H4", "I4:J4"):
        usage.merge_cells(merge_range)
    _font(usage, "A4:J4", Font(bold=True, color=COLORS["teal"], size=15))
    for merge_range in ("A4:B4", "C4:D4", "E4:F4", "G4:H4", "I4:J4"):
        left, top, _, _ = _range_bounds(merge_range)
        usage.cell(row=top, column=left).alignment = Alignment(horizontal="center")
    _number_format(usage, "A4:D4", "#,##0")
    _number_format(usage, "E4:F4", "0.0%")
    _number_format(usage, "G4:H4", "#,##0")
    _number_format(usage, "I4:J4", "#,##0.0")

    usage.merge_cells("A7:J7")
    usage.cell(row=7, column=1).value = "怎么用"
    usage.cell(row=7, column=1).fill = _fill(COLORS["teal"])
    usage.cell(row=7, column=1).font = Font(bold=True, color=COLORS["white"])
    howto_rows = [
        ["1", "找目录", "打开 01_目录索引，按完整路径或主要类型筛选；点击链接直接进入阿里云盘。"],
        ["2", "找文件", "打开 02_明细入口，选择分区并点击“打开明细”；在分区工作簿按文件类型、扩展名或大小筛选。"],
        ["3", "看全盘分布", f"03_类型统计、04_分区统计、05_扩展名统计由 {len(partition_stats)} 份分区缓存聚合，可快速判断资源所在区域。"],
        ["4", "增量更新", "生成器比较 Markdown SHA-256；只重建变化分区及本总入口，未变化分区工作簿直接复用。"],
        ["5", "强制全量", "仅在生成规则或样式整体变化时使用 --force；普通索引刷新不要全量重建。"],
    ]
    for offset, row_values in enumerate(howto_rows):
        row = 8 + offset
        usage.cell(row=row, column=1).value = row_values[0]
        usage.cell(row=row, column=2).value = row_values[1]
        usage.merge_cells(f"C{row}:J{row}")
        usage.cell(row=row, column=3).value = row_values[2]
    for row in range(8, 13):
        for column in range(1, 11):
            cell = usage.cell(row=row, column=column)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        usage.cell(row=row, column=1).fill = _fill(COLORS["pale"])
        usage.cell(row=row, column=1).font = Font(bold=True, color=COLORS["teal"])
        usage.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        usage.cell(row=row, column=2).font = Font(bold=True, color=COLORS["navy"])

    usage.merge_cells("A15:J15")
    usage.cell(row=15, column=1).value = "口径与限制"
    usage.cell(row=15, column=1).fill = _fill(COLORS["teal"])
    usage.cell(row=15, column=1).font = Font(bold=True, color=COLORS["white"])
    coverage_pct = (indexed_files / catalog["totalFiles"] * 100) if catalog["totalFiles"] else 0.0
    note_band(
        usage,
        "A16:J17",
        (
            f"馆藏总览口径：{catalog['totalDirs']:,} 个目录 / {catalog['totalFiles']:,} 个文件 / {catalog['totalSize']}。"
            f"全文检索 Markdown 展开 {indexed_files:,} 个文件，覆盖率 {coverage_pct:.1f}%。"
            "未展开差额主要来自采用聚合索引的分区。"
        ),
        True,
    )
    usage.cell(row=19, column=1).value = "字段"
    usage.cell(row=19, column=2).value = "值"
    usage.cell(row=20, column=1).value = "生成时间"
    usage.cell(row=20, column=2).value = generated_at
    usage.cell(row=21, column=1).value = "云盘入口"
    usage.cell(row=21, column=2).value = "见目录索引与分区统计中的链接"
    usage.cell(row=22, column=1).value = "Bytes/MB"
    usage.cell(row=22, column=2).value = MB
    usage.cell(row=23, column=1).value = "Bytes/GB"
    usage.cell(row=23, column=2).value = GB
    style_header(usage, "A19:B19")
    _number_format(usage, "B22:B23", "#,##0")

    release_metadata = plan.get("releaseMetadata")
    source_start = 32 if release_metadata else 25
    if release_metadata:
        usage.merge_cells("A25:J25")
        usage.cell(row=25, column=1).value = "发布元数据（调用方提供）"
        usage.cell(row=25, column=1).fill = _fill(COLORS["teal"])
        usage.cell(row=25, column=1).font = Font(bold=True, color=COLORS["white"])
        for offset, (key, value) in enumerate(release_metadata_rows(release_metadata)):
            row = 26 + offset
            usage.cell(row=row, column=1).value = key
            usage.cell(row=row, column=2).value = value
        style_header(usage, "A26:A30")
        for row in range(26, 31):
            usage.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

    usage.merge_cells(f"A{source_start}:J{source_start}")
    usage.cell(row=source_start, column=1).value = "数据源"
    usage.cell(row=source_start, column=1).fill = _fill(COLORS["teal"])
    usage.cell(row=source_start, column=1).font = Font(bold=True, color=COLORS["white"])
    source_rows = [[catalog["sourceName"], "馆藏总览"]]
    source_rows.extend([Path(item["source"]).name, "分区全文检索"] for item in plan["partitions"])
    for offset, row_values in enumerate(source_rows):
        row = source_start + 1 + offset
        usage.cell(row=row, column=1).value = row_values[0]
        usage.cell(row=row, column=2).value = row_values[1]
    usage.freeze_panes = "A3"
    set_column_widths(usage, {"A": 24, "B": 42, "C": 22, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15, "I": 15, "J": 15})

    directory_headers = [
        "序号", "分区", "层级", "课程/目录名（点击直达）", "完整路径", "父目录", "主要类型",
        "直接文件数", "子树文件数", "直接大小(Bytes)", "子树大小(Bytes)", "子树大小(GB)",
        "直接子目录数", "文件夹URL", "快捷打开", "来源",
    ]
    last_dir_col = col(len(directory_headers) - 1)
    title_band(directories_sheet, f"A1:{last_dir_col}1", "目录索引")
    note_band(
        directories_sheet,
        f"A2:{last_dir_col}2",
        f"共 {len(directories):,} 行：由 {len(partition_stats)} 份分区缓存反推目录树；可按分区、层级、主要类型、完整路径组合筛选。",
    )
    _write_row(directories_sheet, 4, 0, directory_headers)
    for index, row in enumerate(directories):
        excel_row = index + 5
        _write_row(
            directories_sheet,
            excel_row,
            0,
            [
                index + 1, row["partition"], row["depth"], None, row["path"], row["parent"],
                row["dominantType"], row["directFiles"], row["subtreeFiles"], row["directBytes"],
                row["subtreeBytes"], None, row["childDirCount"], row["url"] or None, None, row["source"],
            ],
        )
        label = excel_text(row["name"])
        directories_sheet.cell(row=excel_row, column=4).value = (
            f'=HYPERLINK(N{excel_row},"{label}")' if row["url"] else f'="{label}"'
        )
        directories_sheet.cell(row=excel_row, column=12).value = f"=K{excel_row}/'00_使用说明'!$B$23"
        directories_sheet.cell(row=excel_row, column=15).value = (
            f'=IF(N{excel_row}="","",HYPERLINK(N{excel_row},"🔗 打开云盘"))'
        )
    if directories:
        add_table(directories_sheet, f"A4:{last_dir_col}{len(directories) + 4}", "DirectoryIndexTable")
        _number_format(directories_sheet, f"A5:C{len(directories) + 4}", "#,##0")
        _number_format(directories_sheet, f"H5:K{len(directories) + 4}", "#,##0")
        _number_format(directories_sheet, f"L5:L{len(directories) + 4}", "#,##0.000")
        _number_format(directories_sheet, f"M5:M{len(directories) + 4}", "#,##0")
        _font(directories_sheet, f"D5:D{len(directories) + 4}", _LINK_FONT)
        _font(directories_sheet, f"O5:O{len(directories) + 4}", _LINK_FONT)
    style_header(directories_sheet, f"A4:{last_dir_col}4")
    directories_sheet.freeze_panes = "C5"
    set_column_widths(
        directories_sheet,
        {
            "A": 9, "B": 20, "C": 8, "D": 38, "E": 58, "F": 45, "G": 16, "H": 13, "I": 13,
            "J": 18, "K": 18, "L": 15, "M": 14, "N": 46, "O": 18, "P": 20,
        },
    )

    entry_headers = [
        "分区", "全盘口径目录数", "全盘口径文件数", "已展开明细", "明细覆盖率", "明细估算大小(Bytes)",
        "明细估算大小(GB)", "全盘口径体量", "分区URL", "打开云盘", "明细工作簿相对路径", "打开明细", "来源Markdown",
    ]
    last_entry_col = col(len(entry_headers) - 1)
    title_band(entry_sheet, f"A1:{last_entry_col}1", "分区文件明细入口")
    note_band(
        entry_sheet,
        f"A2:{last_entry_col}2",
        f"文件明细拆成 {len(partition_stats)} 个独立工作簿。点击“打开明细”进入对应文件；点击“打开云盘”直接进入该分区。相对路径便于整套复制或上传。",
    )
    _write_row(entry_sheet, 4, 0, entry_headers)
    detail_by_partition = {item["partition"]: item["output"] for item in plan["partitions"]}
    for index, row in enumerate(partition_stats):
        excel_row = index + 5
        detail_output = detail_by_partition.get(row["partition"], "")
        relative = friendly_local_path(plan["outputPath"], detail_output) if detail_output else ""
        _write_row(
            entry_sheet,
            excel_row,
            0,
            [
                row["partition"], row["dirs"], row["files"], row["indexedFiles"], None,
                row["indexedBytes"], None, row["volume"], row["url"] or None, None, relative or None, None,
                f"{row['partition']}.md",
            ],
        )
        entry_sheet.cell(row=excel_row, column=5).value = f"=IFERROR(D{excel_row}/C{excel_row},0)"
        entry_sheet.cell(row=excel_row, column=7).value = f"=F{excel_row}/'00_使用说明'!$B$23"
        entry_sheet.cell(row=excel_row, column=10).value = f'=HYPERLINK(I{excel_row},"🔗 打开云盘")'
        entry_sheet.cell(row=excel_row, column=12).value = (
            f'=IF(K{excel_row}="","",HYPERLINK(K{excel_row},"📄 打开明细"))'
        )
    style_header(entry_sheet, f"A4:{last_entry_col}4")
    add_table(entry_sheet, f"A4:{last_entry_col}{partition_end}", "DetailEntryTable")
    entry_sheet.freeze_panes = "A5"
    _number_format(entry_sheet, f"B5:D{partition_end}", "#,##0")
    _number_format(entry_sheet, f"E5:E{partition_end}", "0.0%")
    _number_format(entry_sheet, f"F5:F{partition_end}", "#,##0")
    _number_format(entry_sheet, f"G5:G{partition_end}", "#,##0.00")
    _font(entry_sheet, f"J5:J{partition_end}", _LINK_FONT)
    _font(entry_sheet, f"L5:L{partition_end}", _LINK_FONT)
    set_column_widths(
        entry_sheet,
        {"A": 22, "B": 18, "C": 18, "D": 16, "E": 14, "F": 22, "G": 18, "H": 16, "I": 48, "J": 16, "K": 48, "L": 16, "M": 22},
    )

    type_headers = ["文件类型", "文件数", "明细占比", "估算大小(Bytes)", "估算大小(GB)", "平均大小(MB)", "说明"]
    title_band(type_sheet, "A1:G1", "按文件类型统计")
    note_band(
        type_sheet,
        "A2:G2",
        f"由 {len(partition_stats)} 份分区缓存聚合；文件数和体量是生成器实算值，占比、GB 与平均大小使用工作表公式。",
    )
    _write_row(type_sheet, 4, 0, type_headers)
    for index, row in enumerate(type_stats):
        excel_row = index + 5
        _write_row(type_sheet, excel_row, 0, [row["type"], row["count"], None, row["bytes"], None, None, TYPE_NOTES.get(row["type"], "")])
        type_sheet.cell(row=excel_row, column=3).value = f"=B{excel_row}/'00_使用说明'!$C$4"
        type_sheet.cell(row=excel_row, column=5).value = f"=D{excel_row}/'00_使用说明'!$B$23"
        type_sheet.cell(row=excel_row, column=6).value = f"=IFERROR(D{excel_row}/B{excel_row}/'00_使用说明'!$B$22,0)"
    style_header(type_sheet, "A4:G4")
    add_table(type_sheet, f"A4:G{len(type_stats) + 4}", "TypeSummaryTable")
    type_sheet.freeze_panes = "A5"
    _number_format(type_sheet, f"B5:B{len(type_stats) + 4}", "#,##0")
    _number_format(type_sheet, f"C5:C{len(type_stats) + 4}", "0.0%")
    _number_format(type_sheet, f"D5:D{len(type_stats) + 4}", "#,##0")
    _number_format(type_sheet, f"E5:F{len(type_stats) + 4}", "#,##0.00")
    set_column_widths(type_sheet, {"A": 20, "B": 14, "C": 14, "D": 20, "E": 16, "F": 16, "G": 38})

    partition_headers = [
        "分区", "全盘口径目录数", "全盘口径文件数", "Markdown明细文件数", "明细覆盖率",
        "明细估算大小(Bytes)", "明细估算大小(GB)", "全盘口径体量", "分区URL", "点击直达云盘",
    ]
    title_band(partition_sheet, "A1:J1", "按分区统计与明细覆盖率")
    note_band(
        partition_sheet,
        "A2:J2",
        "全盘口径来自馆藏总览输入；明细口径来自各分区缓存。覆盖率低表示该分区在 Markdown 中采用聚合索引，不代表文件丢失。",
        True,
    )
    _write_row(partition_sheet, 4, 0, partition_headers)
    for index, row in enumerate(partition_stats):
        excel_row = index + 5
        _write_row(
            partition_sheet, excel_row, 0,
            [row["partition"], row["dirs"], row["files"], row["indexedFiles"], None, row["indexedBytes"], None, row["volume"], row["url"] or None, None],
        )
        partition_sheet.cell(row=excel_row, column=5).value = f"=IFERROR(D{excel_row}/C{excel_row},0)"
        partition_sheet.cell(row=excel_row, column=7).value = f"=F{excel_row}/'00_使用说明'!$B$23"
        partition_sheet.cell(row=excel_row, column=10).value = f'=HYPERLINK(I{excel_row},"🔗 打开分区")'
    style_header(partition_sheet, "A4:J4")
    add_table(partition_sheet, f"A4:J{partition_end}", "PartitionSummaryTable")
    partition_sheet.freeze_panes = "A5"
    _number_format(partition_sheet, f"B5:D{partition_end}", "#,##0")
    _number_format(partition_sheet, f"E5:E{partition_end}", "0.0%")
    _number_format(partition_sheet, f"F5:F{partition_end}", "#,##0")
    _number_format(partition_sheet, f"G5:G{partition_end}", "#,##0.00")
    _font(partition_sheet, f"J5:J{partition_end}", _LINK_FONT)
    set_column_widths(partition_sheet, {"A": 22, "B": 18, "C": 18, "D": 20, "E": 16, "F": 22, "G": 18, "H": 16, "I": 48, "J": 18})

    extension_headers = ["扩展名", "文件类型", "文件数", "明细占比", "估算大小(Bytes)", "估算大小(GB)", "筛选提示"]
    title_band(extension_sheet, "A1:G1", "按扩展名统计")
    note_band(
        extension_sheet,
        "A2:G2",
        f"共识别 {len(extension_stats)} 种扩展名；按出现次数排序。文件数和体量来自分区缓存，占比与 GB 使用工作表公式。",
    )
    _write_row(extension_sheet, 4, 0, extension_headers)
    for index, row in enumerate(extension_stats):
        excel_row = index + 5
        hint = "文件名没有后缀" if row["ext"] == "(无扩展名)" else f"在分区明细筛选：{row['ext']}"
        _write_row(extension_sheet, excel_row, 0, [row["ext"], row["type"], row["count"], None, row["bytes"], None, hint])
        extension_sheet.cell(row=excel_row, column=4).value = f"=C{excel_row}/'00_使用说明'!$C$4"
        extension_sheet.cell(row=excel_row, column=6).value = f"=E{excel_row}/'00_使用说明'!$B$23"
    style_header(extension_sheet, "A4:G4")
    add_table(extension_sheet, f"A4:G{len(extension_stats) + 4}", "ExtensionSummaryTable")
    extension_sheet.freeze_panes = "A5"
    _number_format(extension_sheet, f"C5:C{len(extension_stats) + 4}", "#,##0")
    _number_format(extension_sheet, f"D5:D{len(extension_stats) + 4}", "0.0%")
    _number_format(extension_sheet, f"E5:E{len(extension_stats) + 4}", "#,##0")
    _number_format(extension_sheet, f"F5:F{len(extension_stats) + 4}", "#,##0.00")
    set_column_widths(extension_sheet, {"A": 18, "B": 20, "C": 14, "D": 14, "E": 22, "F": 18, "G": 34})

    metrics = {
        "indexedFiles": indexed_files,
        "indexedBytes": indexed_bytes,
        "directories": len(directories),
        "types": len(type_stats),
        "extensions": len(extension_stats),
    }
    return workbook, metrics


def build_partition_workbook(partition_cache: dict[str, Any], generated_at: str) -> tuple[Workbook, dict[str, Any]]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    usage = workbook.create_sheet("00_使用说明")
    files_sheet = workbook.create_sheet("01_文件明细")
    type_sheet = workbook.create_sheet("02_类型统计")
    extension_sheet = workbook.create_sheet("03_扩展名统计")
    for sheet in (usage, files_sheet, type_sheet, extension_sheet):
        prepare_sheet(sheet)

    partition = partition_cache["partition"]
    files = partition_cache["files"]
    source_name = partition_cache["sourceName"]
    type_stats = type_stats_from_files(files)
    extension_stats = extension_stats_from_files(files)
    file_end = len(files) + 4
    category_depth = max([0, *(len(row.get("categories") or []) for row in files)])
    category_headers = [f"分类层级{index + 1}" for index in range(category_depth)]
    type_index0 = 2 + category_depth
    extension_index0 = type_index0 + 1
    size_bytes_index0 = type_index0 + 4
    size_mb_index0 = type_index0 + 5
    folder_url_index0 = type_index0 + 8
    link_index0 = type_index0 + 9

    title_band(usage, "A1:H1", f"{partition} · 文件明细索引")
    note_band(
        usage,
        "A2:H2",
        f"本工作簿只对应一个分区；可按文件类型、扩展名、{category_depth or 0} 层分类和大小组合筛选。"
        "分类列根据真实目录深度动态生成；源 Markdown 未变化时，增量生成器不会重建本文件。",
    )
    _write_row(usage, 4, 0, ["分区", "文件数", "估算体量(GB)", "文件类型数", "扩展名数", "生成时间", "来源Markdown", "缓存策略"])
    style_header(usage, "A4:H4")
    _write_row(usage, 5, 0, [partition, len(files), None, len(type_stats), len(extension_stats), generated_at, source_name, "SHA-256 未变则复用"])
    size_bytes_col = col(size_bytes_index0)
    usage.cell(row=5, column=3).value = f"=SUM('01_文件明细'!${size_bytes_col}$5:${size_bytes_col}${file_end})/$B$9"
    usage.cell(row=5, column=2).number_format = "#,##0"
    usage.cell(row=5, column=3).number_format = "#,##0.00"
    usage.cell(row=5, column=4).number_format = "#,##0"
    usage.cell(row=5, column=5).number_format = "#,##0"
    usage.cell(row=8, column=1).value = "Bytes/MB"
    usage.cell(row=8, column=2).value = MB
    usage.cell(row=9, column=1).value = "Bytes/GB"
    usage.cell(row=9, column=2).value = GB
    usage.merge_cells("A12:H12")
    usage.cell(row=12, column=1).value = "使用建议：先在 02_类型统计或 03_扩展名统计判断资源分布，再回到 01_文件明细筛选；“点击直达云盘”进入文件所在文件夹。"
    usage.cell(row=12, column=1).fill = _fill(COLORS["pale"])
    usage.cell(row=12, column=1).font = Font(color=COLORS["navy"])
    usage.cell(row=12, column=1).alignment = Alignment(wrap_text=True)
    usage.cell(row=15, column=1).value = "源文件"
    usage.cell(row=15, column=2).value = source_name
    usage.cell(row=16, column=1).value = "云盘入口"
    usage.cell(row=16, column=2).value = "见文件明细中的文件夹URL"
    release_metadata = partition_cache.get("releaseMetadata")
    if release_metadata:
        usage.merge_cells("A18:H18")
        usage.cell(row=18, column=1).value = "发布元数据（调用方提供）"
        usage.cell(row=18, column=1).fill = _fill(COLORS["teal"])
        usage.cell(row=18, column=1).font = Font(bold=True, color=COLORS["white"])
        for offset, (key, value) in enumerate(release_metadata_rows(release_metadata)):
            row = 19 + offset
            usage.cell(row=row, column=1).value = key
            usage.cell(row=row, column=2).value = value
        style_header(usage, "A19:A23")
        for row in range(19, 24):
            usage.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    usage.freeze_panes = "A3"
    set_column_widths(usage, {"A": 24, "B": 42, "C": 18, "D": 16, "E": 16, "F": 20, "G": 24, "H": 22})

    file_headers = (
        ["序号", "分区", *category_headers, "文件类型", "扩展名", "文件名", "大小原文", "估算大小(Bytes)",
         "估算大小(MB)", "文件夹路径", "完整文件路径", "文件夹URL", "点击直达云盘", "来源Markdown"]
    )
    last_file_col = col(len(file_headers) - 1)
    title_band(files_sheet, f"A1:{last_file_col}1", f"{partition} · 文件明细（可筛选）")
    note_band(
        files_sheet,
        f"A2:{last_file_col}2",
        f"共 {len(files):,} 个 Markdown 已展开文件。分类层级按实际目录深度动态展开；估算大小由 Markdown 显示值换算；点击链接进入文件所在云盘文件夹。",
    )
    _write_row(files_sheet, 4, 0, file_headers)
    size_mb_col = col(size_mb_index0)
    folder_url_col = col(folder_url_index0)
    link_col = col(link_index0)
    for index, row in enumerate(files):
        excel_row = index + 5
        categories = list(row.get("categories") or [])
        while len(categories) < category_depth:
            categories.append("")
        _write_row(
            files_sheet, excel_row, 0,
            [index + 1, row["partition"], *categories, row["type"], row["ext"], row["name"], row["sizeText"],
             row["sizeBytes"], None, row["folder"], row["fullPath"], row["folderUrl"] or None, None, row["source"]],
        )
        files_sheet.cell(row=excel_row, column=size_mb_index0 + 1).value = f"={size_bytes_col}{excel_row}/'00_使用说明'!$B$8"
        files_sheet.cell(row=excel_row, column=link_index0 + 1).value = (
            f'=IF({folder_url_col}{excel_row}="","",HYPERLINK({folder_url_col}{excel_row},"🔗 打开文件夹"))'
        )
    if files:
        add_table(files_sheet, f"A4:{last_file_col}{file_end}", "FileDetailTable")
        _number_format(files_sheet, f"A5:A{file_end}", "#,##0")
        _number_format(files_sheet, f"{size_bytes_col}5:{size_bytes_col}{file_end}", "#,##0")
        _number_format(files_sheet, f"{size_mb_col}5:{size_mb_col}{file_end}", "#,##0.00")
        _font(files_sheet, f"{link_col}5:{link_col}{file_end}", _LINK_FONT)
    style_header(files_sheet, f"A4:{last_file_col}4")
    files_sheet.freeze_panes = "C5"
    file_widths = {"A": 9, "B": 20}
    for index in range(category_depth):
        file_widths[col(2 + index)] = 28
    file_widths.update(
        {
            col(type_index0): 15,
            col(extension_index0): 13,
            col(type_index0 + 2): 48,
            col(type_index0 + 3): 13,
            col(size_bytes_index0): 18,
            col(size_mb_index0): 16,
            col(type_index0 + 6): 55,
            col(type_index0 + 7): 62,
            col(folder_url_index0): 46,
            col(link_index0): 18,
            col(type_index0 + 10): 22,
        }
    )
    set_column_widths(files_sheet, file_widths)

    type_headers = ["文件类型", "文件数", "明细占比", "估算大小(Bytes)", "估算大小(GB)", "平均大小(MB)", "说明"]
    title_band(type_sheet, "A1:G1", f"{partition} · 按文件类型统计")
    note_band(type_sheet, "A2:G2", "文件数、占比和体量均由本分区文件明细公式计算，可追溯到 01_文件明细。")
    _write_row(type_sheet, 4, 0, type_headers)
    type_col = col(type_index0)
    for index, row in enumerate(type_stats):
        excel_row = index + 5
        _write_row(type_sheet, excel_row, 0, [row["type"], None, None, None, None, None, TYPE_NOTES.get(row["type"], "")])
        type_sheet.cell(row=excel_row, column=2).value = (
            f"=COUNTIF('01_文件明细'!${type_col}$5:${type_col}${file_end},A{excel_row})"
        )
        type_sheet.cell(row=excel_row, column=3).value = f"=B{excel_row}/'00_使用说明'!$B$5"
        type_sheet.cell(row=excel_row, column=4).value = (
            f"=SUMIF('01_文件明细'!${type_col}$5:${type_col}${file_end},A{excel_row},"
            f"'01_文件明细'!${size_bytes_col}$5:${size_bytes_col}${file_end})"
        )
        type_sheet.cell(row=excel_row, column=5).value = f"=D{excel_row}/'00_使用说明'!$B$9"
        type_sheet.cell(row=excel_row, column=6).value = f"=IFERROR(D{excel_row}/B{excel_row}/'00_使用说明'!$B$8,0)"
    style_header(type_sheet, "A4:G4")
    add_table(type_sheet, f"A4:G{len(type_stats) + 4}", "TypeSummaryTable")
    type_sheet.freeze_panes = "A5"
    _number_format(type_sheet, f"B5:B{len(type_stats) + 4}", "#,##0")
    _number_format(type_sheet, f"C5:C{len(type_stats) + 4}", "0.0%")
    _number_format(type_sheet, f"D5:D{len(type_stats) + 4}", "#,##0")
    _number_format(type_sheet, f"E5:F{len(type_stats) + 4}", "#,##0.00")
    set_column_widths(type_sheet, {"A": 20, "B": 14, "C": 14, "D": 20, "E": 16, "F": 16, "G": 38})

    extension_headers = ["扩展名", "文件类型", "文件数", "明细占比", "估算大小(Bytes)", "估算大小(GB)", "筛选提示"]
    title_band(extension_sheet, "A1:G1", f"{partition} · 按扩展名统计")
    note_band(extension_sheet, "A2:G2", f"共识别 {len(extension_stats)} 种扩展名；文件数、占比和体量均由本分区文件明细公式计算。")
    _write_row(extension_sheet, 4, 0, extension_headers)
    extension_col = col(extension_index0)
    for index, row in enumerate(extension_stats):
        excel_row = index + 5
        hint = "文件名没有后缀" if row["ext"] == "(无扩展名)" else f"在文件明细筛选：{row['ext']}"
        _write_row(extension_sheet, excel_row, 0, [row["ext"], row["type"], None, None, None, None, hint])
        extension_sheet.cell(row=excel_row, column=3).value = (
            f"=COUNTIF('01_文件明细'!${extension_col}$5:${extension_col}${file_end},A{excel_row})"
        )
        extension_sheet.cell(row=excel_row, column=4).value = f"=C{excel_row}/'00_使用说明'!$B$5"
        extension_sheet.cell(row=excel_row, column=5).value = (
            f"=SUMIF('01_文件明细'!${extension_col}$5:${extension_col}${file_end},A{excel_row},"
            f"'01_文件明细'!${size_bytes_col}$5:${size_bytes_col}${file_end})"
        )
        extension_sheet.cell(row=excel_row, column=6).value = f"=E{excel_row}/'00_使用说明'!$B$9"
    style_header(extension_sheet, "A4:G4")
    add_table(extension_sheet, f"A4:G{len(extension_stats) + 4}", "ExtensionSummaryTable")
    extension_sheet.freeze_panes = "A5"
    _number_format(extension_sheet, f"C5:C{len(extension_stats) + 4}", "#,##0")
    _number_format(extension_sheet, f"D5:D{len(extension_stats) + 4}", "0.0%")
    _number_format(extension_sheet, f"E5:E{len(extension_stats) + 4}", "#,##0")
    _number_format(extension_sheet, f"F5:F{len(extension_stats) + 4}", "#,##0.00")
    set_column_widths(extension_sheet, {"A": 18, "B": 20, "C": 14, "D": 14, "E": 22, "F": 18, "G": 34})

    metrics = {"partition": partition, "files": len(files), "types": len(type_stats), "extensions": len(extension_stats)}
    return workbook, metrics


def build(plan_path: Path) -> dict[str, Any]:
    """Build all outputs a ``build-plan.json`` requires. Mirrors build_workbooks.mjs's main()."""
    plan = json.loads(Path(plan_path).read_text(encoding="utf-8"))
    aggregate = json.loads(Path(plan["aggregatePath"]).read_text(encoding="utf-8"))
    generated_at = compute_generated_at(plan)
    results = []

    if plan["buildMaster"]:
        workbook, metrics = build_master_workbook(aggregate, plan, generated_at)
        output_path = Path(plan["outputPath"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        results.append({"output": str(output_path), "kind": "master", **metrics})

    for partition_plan in plan["partitions"]:
        if not partition_plan["changed"]:
            continue
        partition_cache = json.loads(Path(partition_plan["cache"]).read_text(encoding="utf-8"))
        workbook, metrics = build_partition_workbook(partition_cache, generated_at)
        output_path = Path(partition_plan["output"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        results.append({"output": str(output_path), "kind": "partition", **metrics})

    payload = {"generatedAt": generated_at, "outputs": results}
    Path(plan["cacheDir"], "build-result.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return payload
