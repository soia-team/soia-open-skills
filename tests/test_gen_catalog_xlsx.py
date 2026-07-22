#!/usr/bin/env python3
"""Tests for the incremental Excel catalog cache and parser."""

from __future__ import annotations

import argparse
import importlib.util
import json
import sys
import os
import subprocess
import tempfile
import unittest
from pathlib import Path
from unittest import mock


REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = REPO_ROOT / "skills" / "soia-pkm-alipan-curator" / "scripts"
SCRIPT = SCRIPTS / "gen_catalog_xlsx.py"
sys.path.insert(0, str(SCRIPTS))

from catalog_xlsx.cache import (  # noqa: E402
    aggregate,
    commit_manifest,
    load_json,
    parse_catalog,
    parse_search_markdown,
    prepare_incremental,
    normalize_release_metadata,
)
import gen_catalog_xlsx  # noqa: E402
from gen_catalog_xlsx import (  # noqa: E402
    cleanup_inspection_sidecars,
    cleanup_stale_partition_outputs,
    resolve_renderer,
)

OPENPYXL_INSTALLED = importlib.util.find_spec("openpyxl") is not None


def search_markdown(partition: str, folder: str, rows: list[tuple[str, str]]) -> str:
    table = "\n".join(f"| {name} | {size} |" for name, size in rows)
    return (
        f"# {partition}\n\n"
        f"> 仅供检索（共 {len(rows):,} 个）\n\n"
        f"## {folder} [🔗打开文件夹](https://example.test/{partition})\n\n"
        "| 文件 | 大小 |\n"
        "|---|---|\n"
        f"{table}\n"
    )


def catalog_markdown() -> str:
    return """# 馆藏总览

> 备份盘 · 全盘 **5 目录 / 3 文件 / 2GB**

| 分区 | 直达 | 目录 | 文件 | 体量 |
|---|---|---:|---:|---:|
| 👶 **10_孩子** | [🔗](https://example.test/child) | 3 | 2 | 1GB |
| 📚 **20_阅读** | [🔗](https://example.test/read) | 2 | 1 | 1GB |

# 👶 [10_孩子](https://example.test/child)

## [10_英语](https://example.test/english)

### [10_课程](https://example.test/course)
"""


class IncrementalCatalogTests(unittest.TestCase):
    @staticmethod
    def fake_node(root: Path) -> Path:
        node = root / "fake-node"
        node.write_text(
            """#!/usr/bin/env python3
import json
import sys
from pathlib import Path

plan = json.loads(Path(sys.argv[sys.argv.index('--plan') + 1]).read_text(encoding='utf-8'))
outputs = [plan['outputPath']] + [item['output'] for item in plan.get('partitions', [])]
for output in outputs:
    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text('fake xlsx', encoding='utf-8')
""",
            encoding="utf-8",
        )
        node.chmod(0o755)
        return node

    def test_cleanup_inspection_sidecars_keeps_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            workbook = Path(temporary) / "catalog.xlsx"
            sidecar = Path(f"{workbook}.inspect.ndjson")
            workbook.write_bytes(b"xlsx")
            sidecar.write_text("debug", encoding="utf-8")
            cleanup_inspection_sidecars([workbook])
            self.assertTrue(workbook.exists())
            self.assertFalse(sidecar.exists())

    def test_release_metadata_rejects_snapshot_after_publish_time(self) -> None:
        with self.assertRaisesRegex(ValueError, "不能晚于"):
            normalize_release_metadata({
                "catalog_release_id": "catalog-20260721.1",
                "index_updated_at": "2026-07-21T09:00:00+08:00",
                "snapshot_at": "2026-07-21T09:01:00+08:00",
                "catalog_schema_version": "2026-07",
                "source_fingerprint": "sha256:abc123",
            })

    def test_cross_sheet_formulas_quote_numeric_sheet_names(self) -> None:
        builder = (SCRIPTS / "catalog_xlsx" / "build_workbooks.mjs").read_text(encoding="utf-8")
        self.assertIn("=SUM('02_明细入口'!", builder)
        self.assertIn("=COUNTA('01_目录索引'!", builder)
        self.assertIn("/'00_使用说明'!", builder)
        self.assertNotIn("=SUM(02_明细入口!", builder)
        self.assertNotIn("=COUNTA(01_目录索引!", builder)

    def test_coverage_formulas_guard_against_zero_total_files_in_mjs_builder(self) -> None:
        """Regression for a real #DIV/0! found via --soffice recalculation against real
        vault data: a partition with 全盘口径文件数=0 (e.g. an empty district) made
        `=D/C` divide by zero in 02_明细入口 and 04_分区统计. Must stay IFERROR-guarded."""
        builder = (SCRIPTS / "catalog_xlsx" / "build_workbooks.mjs").read_text(encoding="utf-8")
        self.assertIn("=IFERROR(D${row}/C${row},0)", builder)
        self.assertNotIn("`=D${row}/C${row}`", builder)

    def test_parser_checks_declared_count(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "10_孩子.md"
            source.write_text(search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB")]), encoding="utf-8")
            parsed = parse_search_markdown(source)
            self.assertEqual(parsed["declared"], 1)
            self.assertEqual(parsed["files"][0]["type"], "视频")
            source.write_text(source.read_text(encoding="utf-8").replace("共 1 个", "共 2 个"), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "与头部声明"):
                parse_search_markdown(source)

    def test_missing_output_dir_uses_home_downloads_default(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            (search_dir / "10_孩子.md").write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB")]),
                encoding="utf-8",
            )
            runtime = root / "runtime"
            artifact_tool_dir = runtime / "node_modules" / "@oai" / "artifact-tool"
            artifact_tool_dir.mkdir(parents=True)
            (artifact_tool_dir / "package.json").write_text('{"name": "@oai/artifact-tool"}', encoding="utf-8")
            node = self.fake_node(root)
            home = root / "home"
            output_dir = home / "Downloads" / "soia-pkm-alipan-curator"
            env = {**os.environ, "HOME": str(home)}
            for name in ("ALIPAN_CURATOR_OUTPUT_DIR", "SOIA_PKM_ALIPAN_CURATOR_CONFIG_FILE"):
                env.pop(name, None)
            result = subprocess.run(
                [
                    "python3",
                    str(SCRIPT),
                    "--catalog", str(catalog),
                    "--search-dir", str(search_dir),
                    "--node", str(node),
                    "--artifact-runtime", str(runtime),
                    "--cache-dir", str(root / "cache"),
                    "--json",
                ],
                check=False,
                capture_output=True,
                text=True,
                env=env,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((output_dir / "00_阿里云盘馆藏总索引.xlsx").is_file())
            self.assertIn(
                f"输出到默认目录 {output_dir}（可用 --output-dir 或 config ALIPAN_CURATOR_OUTPUT_DIR 覆盖）",
                result.stderr,
            )

    def test_environment_output_dir_overrides_home_downloads_default(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            (search_dir / "10_孩子.md").write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB")]),
                encoding="utf-8",
            )
            runtime = root / "runtime"
            artifact_tool_dir = runtime / "node_modules" / "@oai" / "artifact-tool"
            artifact_tool_dir.mkdir(parents=True)
            (artifact_tool_dir / "package.json").write_text('{"name": "@oai/artifact-tool"}', encoding="utf-8")
            node = self.fake_node(root)
            output_dir = root / "configured-output"
            env = {
                **os.environ,
                "HOME": str(root / "home"),
                "ALIPAN_CURATOR_OUTPUT_DIR": str(output_dir),
            }
            env.pop("SOIA_PKM_ALIPAN_CURATOR_CONFIG_FILE", None)
            result = subprocess.run(
                [
                    "python3", str(SCRIPT),
                    "--catalog", str(catalog),
                    "--search-dir", str(search_dir),
                    "--node", str(node),
                    "--artifact-runtime", str(runtime),
                    "--cache-dir", str(root / "cache"),
                    "--json",
                ],
                check=False,
                capture_output=True,
                text=True,
                env=env,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((output_dir / "00_阿里云盘馆藏总索引.xlsx").is_file())
            self.assertNotIn("输出到默认目录", result.stderr)

    def test_parser_preserves_all_classification_levels(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "Library.md"
            source.write_text(
                search_markdown(
                    "Library",
                    "Library/Domain/Topic/Stage/Series",
                    [("a.pdf", "1MB")],
                ),
                encoding="utf-8",
            )

            row = parse_search_markdown(source)["files"][0]

            self.assertEqual(row["categories"], ["Domain", "Topic", "Stage", "Series"])
            self.assertEqual(row["categoryPath"], "Domain/Topic/Stage/Series")
            self.assertEqual(row["categoryDepth"], 4)

    def test_catalog_heading_link_fills_course_parent_url(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            catalog_path = root / "00_馆藏总览.md"
            catalog_path.write_text(catalog_markdown(), encoding="utf-8")
            search_path = root / "10_孩子.md"
            search_path.write_text(
                search_markdown(
                    "10_孩子",
                    "10_孩子/10_英语/10_课程/01_第一课",
                    [("a.mp4", "1MB"), ("b.pdf", "2MB")],
                ),
                encoding="utf-8",
            )
            catalog = parse_catalog(catalog_path)
            result = aggregate(catalog, [parse_search_markdown(search_path)])
            english = next(row for row in result["directories"] if row["path"] == "10_孩子/10_英语")
            course = next(row for row in result["directories"] if row["path"] == "10_孩子/10_英语/10_课程")
            self.assertEqual(english["url"], "https://example.test/english")
            self.assertEqual(course["url"], "https://example.test/course")

    def test_only_changed_partition_is_rebuilt(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            child = search_dir / "10_孩子.md"
            reading = search_dir / "20_阅读.md"
            child.write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB"), ("b.pdf", "2MB")]),
                encoding="utf-8",
            )
            reading.write_text(
                search_markdown("20_阅读", "20_阅读/10_书籍", [("c.epub", "3MB")]),
                encoding="utf-8",
            )
            output = root / "00_阿里云盘馆藏总索引.xlsx"
            cache_dir = root / "cache"

            first = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            self.assertEqual(first["changedPartitions"], ["10_孩子", "20_阅读"])
            aggregate = load_json(Path(first["aggregatePath"]))
            self.assertEqual(aggregate["indexedFiles"], 3)
            self.assertEqual(sum(row["count"] for row in aggregate["typeStats"]), 3)
            output.touch()
            for item in first["partitions"]:
                Path(item["output"]).parent.mkdir(parents=True, exist_ok=True)
                Path(item["output"]).touch()
            commit_manifest(first)

            second = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            self.assertFalse(second["buildMaster"])
            self.assertEqual(second["changedPartitions"], [])

            child.write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a2.mp4", "1MB"), ("b.pdf", "2MB")]),
                encoding="utf-8",
            )
            third = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            self.assertTrue(third["buildMaster"])
            self.assertEqual(third["changedPartitions"], ["10_孩子"])
            self.assertEqual(
                [item["partition"] for item in third["partitions"] if not item["changed"]],
                ["20_阅读"],
            )

    def test_release_metadata_rebuilds_once_and_then_is_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            for partition, folder, name in [
                ("10_孩子", "10_孩子/10_英语", "a.mp4"),
                ("20_阅读", "20_阅读/10_书籍", "b.pdf"),
            ]:
                (search_dir / f"{partition}.md").write_text(
                    search_markdown(partition, folder, [(name, "1MB")]),
                    encoding="utf-8",
                )
            output = root / "00_阿里云盘馆藏总索引.xlsx"
            cache_dir = root / "cache"
            metadata = {
                "catalog_release_id": "catalog-20260721.1",
                "index_updated_at": "2026-07-21T09:30:00+08:00",
                "snapshot_at": "2026-07-21T09:00:00+08:00",
                "catalog_schema_version": "2026-07",
                "source_fingerprint": "sha256:abc123",
            }

            initial = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            output.touch()
            for item in initial["partitions"]:
                Path(item["output"]).parent.mkdir(parents=True, exist_ok=True)
                Path(item["output"]).touch()
            commit_manifest(initial)

            released = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
                release_metadata=metadata,
            )
            self.assertTrue(released["buildMaster"])
            self.assertEqual(released["changedPartitions"], ["10_孩子", "20_阅读"])
            self.assertEqual(released["releaseMetadata"], metadata)
            self.assertIn("releaseMetadataFingerprint", released["nextManifest"])
            for item in released["partitions"]:
                self.assertEqual(load_json(Path(item["cache"]))["releaseMetadata"], metadata)
            commit_manifest(released)

            unchanged = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
                release_metadata=metadata,
            )
            self.assertFalse(unchanged["buildMaster"])
            self.assertEqual(unchanged["changedPartitions"], [])

    def test_release_metadata_requires_all_fields_and_timezone(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            (search_dir / "10_孩子.md").write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB")]),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "snapshot_at"):
                prepare_incremental(
                    catalog_path=catalog,
                    search_dir=search_dir,
                    output_path=root / "catalog.xlsx",
                    cache_dir=root / "cache",
                    release_metadata={
                        "catalog_release_id": "catalog-20260721.1",
                        "index_updated_at": "2026-07-21T09:30:00+08:00",
                        "snapshot_at": "2026-07-21T09:00:00",
                        "catalog_schema_version": "2026-07",
                        "source_fingerprint": "sha256:abc123",
                    },
                )

    def test_removed_partition_rebuilds_master_and_cleans_stale_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            search_dir = root / "search"
            search_dir.mkdir()
            catalog = root / "00_馆藏总览.md"
            catalog.write_text(catalog_markdown(), encoding="utf-8")
            child = search_dir / "10_孩子.md"
            reading = search_dir / "20_阅读.md"
            child.write_text(
                search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB")]),
                encoding="utf-8",
            )
            reading.write_text(
                search_markdown("20_阅读", "20_阅读/10_书籍", [("b.pdf", "2MB")]),
                encoding="utf-8",
            )
            output = root / "00_阿里云盘馆藏总索引.xlsx"
            cache_dir = root / "cache"
            first = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            output.touch()
            for item in first["partitions"]:
                Path(item["output"]).parent.mkdir(parents=True, exist_ok=True)
                Path(item["output"]).touch()
            commit_manifest(first)

            reading.unlink()
            second = prepare_incremental(
                catalog_path=catalog,
                search_dir=search_dir,
                output_path=output,
                cache_dir=cache_dir,
            )
            self.assertTrue(second["buildMaster"])
            self.assertEqual(
                [item["partition"] for item in second["stalePartitions"]],
                ["20_阅读"],
            )
            stale_output = Path(second["stalePartitions"][0]["output"])
            self.assertTrue(stale_output.exists())
            unmanaged_output = stale_output.parent / "旧分区.xlsx"
            unmanaged_output.touch()
            removed = cleanup_stale_partition_outputs(second)
            self.assertEqual(
                removed,
                sorted([str(stale_output.resolve()), str(unmanaged_output.resolve())]),
            )
            self.assertFalse(stale_output.exists())
            self.assertFalse(unmanaged_output.exists())
            self.assertTrue((stale_output.parent / "10_孩子.xlsx").exists())


class RendererSelectionTests(unittest.TestCase):
    """gen_catalog_xlsx.py must not hard-require @oai/artifact-tool: other AIs/hosts
    without access to that bespoke runtime still need to produce a valid catalog Excel
    via the openpyxl fallback. See resolve_renderer()."""

    @staticmethod
    def make_artifact_runtime(root: Path) -> Path:
        runtime = root / "runtime"
        artifact_tool_dir = runtime / "node_modules" / "@oai" / "artifact-tool"
        artifact_tool_dir.mkdir(parents=True)
        (artifact_tool_dir / "package.json").write_text('{"name": "@oai/artifact-tool"}', encoding="utf-8")
        return runtime

    def test_auto_prefers_artifact_tool_when_fully_available(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            runtime = self.make_artifact_runtime(Path(temporary))
            args = argparse.Namespace(renderer="auto", artifact_runtime=runtime, node="python3")
            self.assertEqual(resolve_renderer(args), "artifact-tool")

    def test_auto_falls_back_to_openpyxl_when_artifact_tool_missing(self) -> None:
        args = argparse.Namespace(renderer="auto", artifact_runtime=None, node="python3")
        with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=True):
            self.assertEqual(resolve_renderer(args), "openpyxl")

    def test_auto_falls_back_to_openpyxl_when_node_missing_even_with_runtime_dir(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            runtime = self.make_artifact_runtime(Path(temporary))
            args = argparse.Namespace(renderer="auto", artifact_runtime=runtime, node="definitely-not-a-real-node-binary-xyz")
            with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=True):
                self.assertEqual(resolve_renderer(args), "openpyxl")

    def test_auto_raises_when_neither_backend_available(self) -> None:
        args = argparse.Namespace(renderer="auto", artifact_runtime=None, node="definitely-not-a-real-node-binary-xyz")
        with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=False):
            with self.assertRaisesRegex(SystemExit, "openpyxl"):
                resolve_renderer(args)

    def test_auto_falls_back_to_openpyxl_when_artifact_tool_dir_exists_but_is_stale(self) -> None:
        """A directory named node_modules/@oai/artifact-tool existing is not proof the
        package actually works -- a stale/partial provisioning must not make 'auto'
        commit to a backend that will then hard-fail instead of falling back."""
        with tempfile.TemporaryDirectory() as temporary:
            runtime = Path(temporary) / "runtime"
            (runtime / "node_modules" / "@oai" / "artifact-tool").mkdir(parents=True)  # no package.json
            args = argparse.Namespace(renderer="auto", artifact_runtime=runtime, node="python3")
            with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=True):
                self.assertEqual(resolve_renderer(args), "openpyxl")

    def test_explicit_artifact_tool_requires_runtime(self) -> None:
        args = argparse.Namespace(renderer="artifact-tool", artifact_runtime=None, node="python3")
        with self.assertRaisesRegex(SystemExit, "artifact-runtime"):
            resolve_renderer(args)

    def test_explicit_artifact_tool_requires_package_present(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            empty_runtime = Path(temporary) / "runtime"
            empty_runtime.mkdir()
            args = argparse.Namespace(renderer="artifact-tool", artifact_runtime=empty_runtime, node="python3")
            with self.assertRaisesRegex(SystemExit, "artifact-tool"):
                resolve_renderer(args)

    def test_explicit_openpyxl_requires_package_installed(self) -> None:
        args = argparse.Namespace(renderer="openpyxl", artifact_runtime=None, node="python3")
        with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=False):
            with self.assertRaisesRegex(SystemExit, "openpyxl 不可用"):
                resolve_renderer(args)

    def test_explicit_openpyxl_ignores_available_artifact_tool(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            runtime = self.make_artifact_runtime(Path(temporary))
            args = argparse.Namespace(renderer="openpyxl", artifact_runtime=runtime, node="python3")
            with mock.patch.object(gen_catalog_xlsx, "openpyxl_available", return_value=True):
                self.assertEqual(resolve_renderer(args), "openpyxl")


@unittest.skipUnless(OPENPYXL_INSTALLED, "openpyxl not installed in this test environment")
class OpenpyxlFallbackIntegrationTests(unittest.TestCase):
    """End-to-end coverage for the platform-agnostic renderer: no @oai/artifact-tool,
    no Node.js — just gen_catalog_xlsx.py + openpyxl, so any AI/host can run it."""

    def build(self, root: Path, *, verify: bool = False, extra_args: list[str] | None = None) -> dict:
        search_dir = root / "search"
        search_dir.mkdir(exist_ok=True)
        catalog = root / "00_馆藏总览.md"
        catalog.write_text(catalog_markdown(), encoding="utf-8")
        (search_dir / "10_孩子.md").write_text(
            search_markdown("10_孩子", "10_孩子/10_英语", [("a.mp4", "1MB"), ("b.pdf", "2MB")]),
            encoding="utf-8",
        )
        (search_dir / "20_阅读.md").write_text(
            search_markdown("20_阅读", "20_阅读/10_书籍", [("c.epub", "3MB")]),
            encoding="utf-8",
        )
        command = [
            "python3", str(SCRIPT),
            "--catalog", str(catalog),
            "--search-dir", str(search_dir),
            "--output-dir", str(root / "output"),
            "--cache-dir", str(root / "cache"),
            "--renderer", "openpyxl",
            "--json",
        ]
        if verify:
            command.append("--verify")
        if extra_args:
            command.extend(extra_args)
        env = {**os.environ}
        for name in ("ALIPAN_CURATOR_OUTPUT_DIR", "SOIA_PKM_ALIPAN_CURATOR_CONFIG_FILE", "SOIA_ARTIFACT_RUNTIME"):
            env.pop(name, None)
        result = subprocess.run(command, check=False, capture_output=True, text=True, env=env)
        self.assertEqual(result.returncode, 0, result.stderr)
        return json.loads(result.stdout)

    def test_openpyxl_renderer_produces_valid_workbooks_with_real_data(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            payload = self.build(root)
            self.assertEqual(payload["status"], "updated")
            self.assertEqual(payload["renderer"], "openpyxl")

            master_path = Path(payload["master"])
            self.assertTrue(master_path.is_file())
            master = load_workbook(master_path)
            self.assertEqual(
                master.sheetnames,
                ["00_使用说明", "01_目录索引", "02_明细入口", "03_类型统计", "04_分区统计", "05_扩展名统计"],
            )
            usage = master["00_使用说明"]
            self.assertEqual(usage["A4"].value, 3)
            entry = master["02_明细入口"]
            self.assertEqual(entry["A5"].value, "10_孩子")
            self.assertEqual(entry["D5"].value, 2)

            detail_dir = Path(payload["detailDir"])
            partition = load_workbook(detail_dir / "10_孩子.xlsx")
            self.assertEqual(
                partition.sheetnames, ["00_使用说明", "01_文件明细", "02_类型统计", "03_扩展名统计"]
            )
            files_sheet = partition["01_文件明细"]
            header_row = {cell.value: cell.column for cell in files_sheet[4]}
            self.assertEqual(files_sheet["A5"].value, 1)
            self.assertEqual(files_sheet.cell(row=5, column=header_row["文件名"]).value, "a.mp4")
            # HYPERLINK formula string, not just a bare url — must stay clickable.
            link_cell = files_sheet.cell(row=5, column=header_row["点击直达云盘"])
            self.assertTrue(str(link_cell.value).startswith("=IF("))

    def test_untrusted_filenames_are_never_written_as_live_formulas(self) -> None:
        """Alipan file/folder names are attacker-controlled (any uploader can name a
        file '=cmd|...'). openpyxl auto-promotes a bare string starting with '=' into
        a live formula cell (data_type 'f') unless explicitly forced back to 's' --
        this is a real, previously-unguarded formula-injection hole found via
        adversarial review and reproduced end-to-end against build_partition_workbook."""
        from openpyxl import load_workbook
        from catalog_xlsx.build_workbooks_fallback import build_partition_workbook

        payload = '=cmd|"/c calc"!A0'
        partition_cache = {
            "partition": "10_孩子",
            "sourceName": "10_孩子.md",
            "files": [
                {
                    "partition": "10_孩子", "categories": ["10_英语"], "type": "视频", "ext": "mp4",
                    "name": f"{payload}.mp4", "sizeText": "1MB", "sizeBytes": 1048576,
                    "folder": "10_孩子/10_英语", "fullPath": f"10_孩子/10_英语/{payload}.mp4",
                    "folderUrl": "https://example.test/10_孩子", "source": "10_孩子.md",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as temporary:
            workbook, _ = build_partition_workbook(partition_cache, "2026-07-22 00:00")
            output = Path(temporary) / "partition.xlsx"
            workbook.save(output)
            reloaded = load_workbook(output)
            files_sheet = reloaded["01_文件明细"]
            header_row = {cell.value: cell.column for cell in files_sheet[4]}
            name_cell = files_sheet.cell(row=5, column=header_row["文件名"])
            self.assertNotEqual(name_cell.data_type, "f", "filename was silently promoted to a live formula")
            self.assertEqual(name_cell.value, f"{payload}.mp4")

    def test_entry_sheet_survives_and_blanks_link_for_partition_without_search_file(self) -> None:
        """Real-world shape: 80_待探索资源 appears in the catalog's partition table but
        has no matching _全文检索/*.md (it's currently empty), so it has no entry in
        plan['partitions']. Must not crash and must not emit a dead HYPERLINK formula."""
        from catalog_xlsx.build_workbooks_fallback import build_master_workbook

        catalog = {
            "sourceName": "00_馆藏总览.md", "totalDirs": 4, "totalFiles": 1, "totalSize": "1MB",
            "partitions": [
                {"partition": "10_孩子", "url": "https://example.test/10", "dirs": 3, "files": 1, "volume": "1MB"},
                {"partition": "80_待探索资源", "url": "https://example.test/80", "dirs": 1, "files": 0, "volume": "0B"},
            ],
            "headingLinks": [],
        }
        aggregate = {
            "catalog": catalog, "indexedFiles": 1, "indexedBytes": 1048576,
            "directories": [], "typeStats": [], "extensionStats": [],
            "partitionStats": [
                {"partition": "10_孩子", "url": "https://example.test/10", "dirs": 3, "files": 1, "volume": "1MB", "indexedFiles": 1, "indexedBytes": 1048576},
                {"partition": "80_待探索资源", "url": "https://example.test/80", "dirs": 1, "files": 0, "volume": "0B", "indexedFiles": 0, "indexedBytes": 0},
            ],
        }
        # 80_待探索资源 has no entry here -- mirrors prepare_incremental() when there is
        # no corresponding _全文检索/80_待探索资源.md on disk.
        plan = {
            "outputPath": "/tmp/does-not-matter.xlsx",
            "partitions": [
                {"partition": "10_孩子", "output": "/tmp/does-not-matter-分区明细/10_孩子.xlsx", "source": "/tmp/search/10_孩子.md"}
            ],
            "releaseMetadata": None,
        }
        workbook, _ = build_master_workbook(aggregate, plan, "2026-07-22 00:00")
        entry_sheet = workbook["02_明细入口"]
        # 80_待探索资源 is the 2nd partition -> row 6.
        self.assertIn(entry_sheet["K6"].value, (None, ""))
        link_formula = entry_sheet["L6"].value
        self.assertIn('IF(K6=""', link_formula)

    def test_mjs_friendly_local_path_guards_against_missing_detail_path(self) -> None:
        style = (SCRIPTS / "catalog_xlsx" / "workbook_style.mjs").read_text(encoding="utf-8")
        self.assertIn("if (!detailPath) return", style)

    def test_mjs_entry_sheet_open_detail_link_is_guarded(self) -> None:
        builder = (SCRIPTS / "catalog_xlsx" / "build_workbooks.mjs").read_text(encoding="utf-8")
        self.assertIn('IF(K${row}=""', builder)

    def test_openpyxl_renderer_second_run_is_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = self.build(root)
            self.assertEqual(first["status"], "updated")
            second = self.build(root)
            self.assertEqual(second["status"], "unchanged")
            self.assertEqual(second["rebuilt"], [])

    def test_openpyxl_renderer_without_soffice_notes_qa_gap_instead_of_silently_passing(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            payload = self.build(root, verify=True)
            self.assertIsNotNone(payload["qa"])
            self.assertIsNone(payload["qa"].get("formulaErrorCount"))
            self.assertIn("soffice", payload["qa"]["note"])

    def test_coverage_formulas_guard_against_zero_total_files_partition(self) -> None:
        """A partition with 全盘口径文件数=0 (e.g. an empty district like 80_待探索资源)
        must not produce a bare D/C division — that evaluates to #DIV/0! in real Excel/
        LibreOffice, caught via --soffice recalculation against real vault data."""
        from catalog_xlsx.build_workbooks_fallback import build_master_workbook

        catalog = {
            "sourceName": "00_馆藏总览.md",
            "totalDirs": 4,
            "totalFiles": 1,
            "totalSize": "1MB",
            "partitions": [
                {"partition": "10_孩子", "url": "https://example.test/10", "dirs": 3, "files": 1, "volume": "1MB"},
                {"partition": "80_待探索资源", "url": "https://example.test/80", "dirs": 1, "files": 0, "volume": "0B"},
            ],
            "headingLinks": [],
        }
        aggregate = {
            "catalog": catalog,
            "indexedFiles": 1,
            "indexedBytes": 1048576,
            "directories": [],
            "typeStats": [],
            "extensionStats": [],
            "partitionStats": [
                {"partition": "10_孩子", "url": "https://example.test/10", "dirs": 3, "files": 1, "volume": "1MB", "indexedFiles": 1, "indexedBytes": 1048576},
                {"partition": "80_待探索资源", "url": "https://example.test/80", "dirs": 1, "files": 0, "volume": "0B", "indexedFiles": 0, "indexedBytes": 0},
            ],
        }
        plan = {"outputPath": "/tmp/does-not-matter.xlsx", "partitions": [], "releaseMetadata": None}
        workbook, _ = build_master_workbook(aggregate, plan, "2026-07-22 00:00")
        # 80_待探索资源 is the 2nd partition -> row 6 (data starts at row 5).
        entry_formula = workbook["02_明细入口"]["E6"].value
        partition_formula = workbook["04_分区统计"]["E6"].value
        self.assertIn("IFERROR", entry_formula)
        self.assertIn("IFERROR", partition_formula)


if __name__ == "__main__":
    unittest.main()
